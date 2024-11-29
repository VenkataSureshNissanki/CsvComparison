import csv
import datetime
import os
import openpyxl
import pandas as pd
import numpy as np

from csv_diff import compare, load_csv


def get_csv_differences(path1, path2, key):
    pd.set_option('display.max_rows', None)
    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', None)
    pd.option_context('display.colheader_justify', 'right')
    # reading the excel file for dynamically to pass file paths based on lists

    rand = datetime.datetime.now().strftime('%m_%d_%Y_%H_%M_%S')
    # files names for the dev and prd files after the setting index for "primary key"  of the respective file
    dev_file = "test" + os.sep + "subFiles" + os.sep + "dev_file_" + str(rand) + ".csv"
    prd_file = "test" + os.sep + "subFiles" + os.sep + "prd_file_" + str(rand) + ".csv"
    final_merge_files = "test" + os.sep + "subFiles" + os.sep + "merge_files_" + str(rand) + ".csv"

    # file path 1
    file1 = path1
    # file path 2
    file2 = path2

    # common key on both files
    common_key_on_both = key
    print(common_key_on_both)

    df1 = pd.read_csv(file1, parse_dates=True, infer_datetime_format=True, encoding='utf8').fillna("empty")
    df2 = pd.read_csv(file2, parse_dates=True, infer_datetime_format=True, encoding='utf8').fillna("empty")
    di = dict(get_mismatched_dtypes(df1, df2))
    print(di)
    mer_df1 = df1
    mer_df2 = df2
    df1 = df1.set_index(common_key_on_both)
    df2 = df2.set_index(common_key_on_both)
    # df1 = df1.reset_index()
    # df2 = df2.reset_index()
    df1 = df1.sort_values(common_key_on_both, ascending=True)
    df2 = df2.sort_values(common_key_on_both, ascending=True)
    df1.to_csv(dev_file, encoding='utf8')  # , encoding='utf8'
    df2.to_csv(prd_file, encoding='utf8')  # , encoding='utf8'
    # _____________________________ new code______________________________
    # mer_df1 = mer_df1.set_index(common_key_on_both)
    # mer_df2 = mer_df2.set_index(common_key_on_both)
    # d = {"left_only": "Dev_Only", "right_only": "Prod_Only",
    #      "both": "Present in Dev and Prod"}
    # res = mer_df1.merge(mer_df2, indicator=True, on=common_key_on_both, how='outer').loc[
    #     lambda v: v['_merge'] != 'both']
    # merg = res['_merge'].map(d)
    # res['_merge'] = merg
    # res = res.reset_index()
    # res = res.set_index("_merge")
    # res = res.reset_index()
    # res = res.rename({'_merge': 'dev_or_Prd_only'}, axis=1)
    #
    # print(type(res))
    # res.to_csv("diff_1.csv")
    # _____________________________ new code______________________________
    diff = compare(
        load_csv(open(dev_file, encoding='utf8'), key=common_key_on_both),
        load_csv(open(prd_file, encoding='utf8'), key=common_key_on_both)
    )
    # print(type(diff))
    # print(diff)
    # print(diff.keys())
    changes = len(diff['changed'])
    # if len(diff['changed']) > 0:
    df = pd.DataFrame.from_dict(diff, orient='index')
    # df.to_csv("dif_0.csv")
    df = df.transpose()
    if changes > 0:
        dfg = df["changed"]
    else:
        dfg = pd.DataFrame()
    # dfg.to_csv("diff_2.csv")
    # print(dfg)
    lis = []
    ls = []
    for i in range(0, dfg.shape[0]):
        for k, v in dfg[i].items():
            if k == "key":
                ls.append(list_to_str2(common_key_on_both, v))
            if k == "changes":
                lis4 = [key + " --> " + "SRC: " + val[0] + " " + "TARG: " + val[1] for key, val in v.items()]
                lis.append(lis4)
    # lis.insert(0,ls)
    # print(ls)
    # print(lis)
    print(len(lis))
    # print(res.shape[0]) or res.shape[0] > 0
    if len(lis) > 0 :
        p_f = 'Failed'
    else:
        p_f = 'Passed'
    final_csv_with_differences = "test" + os.sep + "Final_Csv_Files" + os.sep + "csv_with_differences_" + \
                                 get_basename(path1) + "_" + p_f + "_" + str(rand) + ".csv"
    final_report_html_file = "test" + os.sep + "Final_Html_Files" + os.sep + \
                             get_basename(path1) + "_" + p_f + "_" + str(rand) + ".html"
    cols = ["key_on_both", "Differences on both"]
    dh1 = pd.DataFrame(list(zip(ls, lis)), columns=cols)
    # print(dh1.to_string())
    # if html_report_y_or_n : fpdf and outlook
    dh1.to_csv(final_csv_with_differences, index=False)
    header_dev = "Src File"
    header_prod = "Tar File"
    with open(final_report_html_file, 'w', encoding='utf8') as _file:
        _file.write('<center>'
                    '<table border="1" class="Final">'
                    + '<tbody>'
                    + '<tr><td><h4>' + header_dev + '</h4></td></tr><tr><td>' + file1 + '</td></tr>'
                    + '<tr><td><h4>' + header_prod + '</h4></td></tr><tr><td>' + file2 + '</td></tr>'
                    + '<tr><th><h4>' + "Differences in Both Csv Files" + '<h4></th></tr><tr><td>'
                    + dh1.to_html(index=False, border=2, justify="center") + '</td></tr>'
                    + '<tr><td><h4>' + " Differences of Dev_Only and Prod_Only" + ' </h4></th></tr><tr><td>'
                    # + res.to_html(index=False, border=2, justify="center") + '</td></tr>'
                    + '<tbody>'
                    + '</table>'
                    + '</center>')


def list_to_str(val):
    li1 = ["Src: " + str(val[0]), "Tar: " + str(val[1])]
    return li1


def list_to_str2(key, val):
    trd_id = key + ": " + str(val)
    return trd_id


def dataframe_difference(df1, df2, which=None):
    """Find rows which are different between two DataFrames."""
    comparison_df = df1.merge(
        df2,
        indicator=True,
        how='outer'
    )
    if which is None:
        diff_df = comparison_df[comparison_df['_merge'] != 'both']
    else:
        diff_df = comparison_df[comparison_df['_merge'] == which]
    # diff_df.to_csv('data/diff.csv')
    return diff_df


def read_excel(xl_filename, sheet_name):
    path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    filename = path + os.sep + xl_filename
    # print(filename)
    wb_obj = openpyxl.load_workbook(filename)
    sheet_obj = wb_obj[sheet_name]
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row

    ls2 = []
    for i in range(2, max_row + 1):
        ls = []
        for j in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=i, column=j)
            # print(cell_obj.value)
            # print(type(cell_obj.value))
            ls.append(cell_obj.value)
        ls2.append(ls)
    # print(ls2)
    return ls2


def read_list(list_num, file_name, sheet_name):
    ls_of_ls = read_excel(file_name, sheet_name)
    # print(ls_of_ls)
    return ls_of_ls[list_num]


def get_mismatched_dtypes(df1, df2):
    df1_types = set(df1.dtypes.items())
    df2_types = set(df2.dtypes.items())
    for column_name, df1_type in df1_types - df2_types:
        yield column_name, (df1_type, df2.dtypes[column_name])


def get_cols_uppercase(list_to_change, u_l):
    if u_l:
        list_of_columns_converted = [str(x).upper().strip() for x in list_to_change.split(',') if x]
    else:
        list_of_columns_converted = [str(x).strip() for x in list_to_change.split(',') if x]
    # print(list_of_columns_converted)
    return list_of_columns_converted


def get_basename(file_name_from_path):
    return os.path.basename(file_name_from_path)


def get_file_path(file):
    path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    final_path = os.path.join(path, file)
    return final_path


def get_file_names(file):
    # path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # final_path = os.path.join(path, file)
    # print(final_path)
    ls = read_list(1)
    # print(os.path.splitext(os.path.basename(ls[0]))[0])
    rand = datetime.datetime.now().strftime('%m_%d_%Y_%H_%M_%S')
    return [os.path.splitext(os.path.basename(ls[0]))[0] + "_" + str(rand) + ".csv",
            os.path.splitext(os.path.basename(ls[1]))[0] + "_" + str(rand) + ".csv"]


def csv_generator_from_text_file(path, fmt, cols):
    # td2csv --> text file with delimeter file to csv
    pd.set_option('display.max_rows', None)
    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', None)
    pd.option_context('display.colheader_justify', 'right')
    # time_stamp = str(pd.datetime.now())
    # print(type(time_stamp))
    # time_stamp_1 = time_stamp.split()[1]
    time_stamp_1 = datetime.datetime.now().strftime('%m_%d_%Y_%H_%M_%S')
    print(time_stamp_1)
    before_name = "test" + os.sep + "converted_files" + os.sep + "delmeter_" + str(time_stamp_1) + ".csv"
    final_name = "test" + os.sep + "converted_files" + os.sep + get_basename(path) + "_" + str(time_stamp_1) + ".csv"
    print(final_name)
    # print(path)
    path_of_gen_exl_file = ''
    excel_file_name = "test" + os.sep + "converted_files" + os.sep + "conv_exl2csv_" + get_basename(path) + "_" + str(
        time_stamp_1) + ".csv"
    if fmt == 'xlsx' or fmt == 'xls':
        if fmt == 'xls':
            read_file = pd.read_excel(path, engine='xlrd')
        else:
            read_file = pd.read_excel(path)
        read_file.to_csv(excel_file_name, index=None, header=True)
        path_of_gen_exl_file = get_file_path(excel_file_name)
        print(path_of_gen_exl_file)
        dh1 = pd.read_csv(path_of_gen_exl_file)
    elif fmt == 'txt':
        dh1 = pd.read_csv(path)
        # ----- new change
        dh1.to_csv("test" + os.sep + "subFiles" + os.sep + "file1.csv", index=False)
        lis = []
        with open("test" + os.sep + "subFiles" + os.sep + "file1.csv", "r") as csvfile:
            reader = csv.reader(csvfile, delimiter='|')
            for row in reader:
                lis.append(row)
        # print(lis[0])
        ls = lis[0]
        lis.pop(0)
        dh1 = pd.DataFrame(lis, columns=ls)
        # dh1 = dh1.reset_index().rename(columns={'index': 'Primary_Key'})
        # quotechar='"',encoding='ISO-8859-1', na_filter=False, low_memory=False
        dh1.to_csv(before_name, index=False)
    else:  # ,encoding='ISO-8859-1'  quotechar='"'
        dh1 = pd.read_csv(path, encoding='utf8')
    # ----- new change


    dh1['Key_to_match'] = dh1[cols].apply(lambda row: '_'.join(row.values.astype(str)), axis=1)
    dh1 = dh1.set_index("Key_to_match")
    dh1 = dh1.reset_index()
    dh1 = dh1.sort_values('Key_to_match', ascending=True)
    dh1.to_csv(final_name, index=False)
    path_of_gen_file = get_file_path(final_name)
    print(path_of_gen_file)
    return path_of_gen_file


def get_lmrs_csv_differences(path1, path2, key, dt):
    pd.set_option('display.max_rows', None)
    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', None)
    pd.option_context('display.colheader_justify', 'right')
    # reading the excel file for dynamically to pass file paths based on lists

    rand = datetime.datetime.now().strftime('%m_%d_%Y_%H_%M_%S')
    # files names for the dev and prd files after the setting index for "primary key"  of the respective file
    dev_file = "test" + os.sep + "subFiles" + os.sep + "dev_file_" + str(rand) + ".csv"
    prd_file = "test" + os.sep + "subFiles" + os.sep + "prd_file_" + str(rand) + ".csv"
    final_merge_files = "test" + os.sep + "subFiles" + os.sep + "merge_files_" + str(rand) + ".csv"

    # file path 1
    file1 = path1
    # file path 2
    file2 = path2

    # common key on both files
    common_key_on_both = key
    print(common_key_on_both)

    df1 = pd.read_csv(file1, parse_dates=True, infer_datetime_format=True, encoding='utf8').fillna("empty")
    df2 = pd.read_csv(file2, parse_dates=True, infer_datetime_format=True, encoding='utf8').fillna("empty")
    di = dict(get_mismatched_dtypes(df1, df2))
    print(di)
    mer_df1 = df1
    mer_df2 = df2
    df1 = df1.set_index(common_key_on_both)
    df2 = df2.set_index(common_key_on_both)
    # df1 = df1.reset_index()
    # df2 = df2.reset_index()
    df1 = df1.sort_values(common_key_on_both, ascending=True)
    df2 = df2.sort_values(common_key_on_both, ascending=True)
    df1.to_csv(dev_file, encoding='utf8')  # , encoding='utf8'
    df2.to_csv(prd_file, encoding='utf8')  # , encoding='utf8'
    shape_of_file = df1.shape
    # _____________________________ new code______________________________
    # mer_df1 = mer_df1.set_index(common_key_on_both)
    # mer_df2 = mer_df2.set_index(common_key_on_both)
    # d = {"left_only": "Dev_Only", "right_only": "Prod_Only",
    #      "both": "Present in Dev and Prod"}
    # res = mer_df1.merge(mer_df2, indicator=True, on=common_key_on_both, how='outer').loc[
    #     lambda v: v['_merge'] != 'both']
    # merg = res['_merge'].map(d)
    # res['_merge'] = merg
    # res = res.reset_index()
    # res = res.set_index("_merge")
    # res = res.reset_index()
    # res = res.rename({'_merge': 'dev_or_Prd_only'}, axis=1)
    #
    # print(type(res))
    # res.to_csv("diff_1.csv")
    # _____________________________ new code______________________________
    diff = compare(
        load_csv(open(dev_file, encoding='utf8'), key=common_key_on_both),
        load_csv(open(prd_file, encoding='utf8'), key=common_key_on_both)
    )
    # print(type(diff))
    # print(diff)
    # print(diff.keys())
    changes = len(diff['changed'])
    # if len(diff['changed']) > 0:
    df = pd.DataFrame.from_dict(diff, orient='index')
    # df.to_csv("dif_0.csv")
    df = df.transpose()
    if changes > 0:
        dfg = df["changed"]
    else:
        dfg = pd.DataFrame()
    # dfg.to_csv("diff_2.csv")
    # print(dfg)
    lis = []
    ls = []
    for i in range(0, dfg.shape[0]):
        for k, v in dfg[i].items():
            if k == "key":
                ls.append(list_to_str2(common_key_on_both, v))
            if k == "changes":
                lis4 = [key + " --> " + "SRC: " + val[0] + " " + "TARG: " + val[1] for key, val in v.items()]
                lis.append(lis4)
    # lis.insert(0,ls)
    # print(ls)
    # print(lis)
    print(len(lis))
    count = len(lis)
    # print(res.shape[0]) or res.shape[0] > 0
    if len(lis) > 0 :
        p_f = 'Failed'
    else:
        p_f = 'Passed'
    final_csv_with_differences = "test" + os.sep + "Final_Csv_Files" + os.sep + "csv_with_differences_" + \
                                 get_basename(path1) + "_" + p_f + "_" + str(rand) + ".csv"
    final_report_html_file = "test" + os.sep + "Final_Html_Files" + os.sep + \
                             get_basename(path1) + "_" + p_f + "_" + str(rand) + ".html"
    cols = ["key_on_both", "Differences on both"]
    dh1 = pd.DataFrame(list(zip(ls, lis)), columns=cols)
    # print(dh1.to_string())
    # if html_report_y_or_n : fpdf and outlook
    dh1.to_csv(final_csv_with_differences, index=False)
    header_dev = "Src File"
    header_prod = "Tar File"
    bus_dt = "Business_Date"
    tot_diff = "Total Differnces_count"
    sh_of_file = "Rows , Columns"
    with open(final_report_html_file, 'w', encoding='utf8') as _file:
        _file.write('<center>'
                    '<table border="1" class="Final">'
                    + '<tbody>'
                    + '<tr><td><h4>' + bus_dt + '</h4></td></tr><tr><td>' + str(dt) + '</td></tr>'
                    + '<tr><td><h4>' + tot_diff + '</h4></td></tr><tr><td>' + str(count) + '</td></tr>'
                    + '<tr><td><h4>' + sh_of_file + '</h4></td></tr><tr><td>' + str(shape_of_file) + '</td></tr>'
                    + '<tr><td><h4>' + header_dev + '</h4></td></tr><tr><td>' + file1 + '</td></tr>'
                    + '<tr><td><h4>' + header_prod + '</h4></td></tr><tr><td>' + file2 + '</td></tr>'
                    + '<tr><th><h4>' + "Differences in Both Csv Files" + '<h4></th></tr><tr><td>'
                    + dh1.to_html(index=False, border=2, justify="center") + '</td></tr>'
                    + '<tr><td><h4>' + " Differences of Dev_Only and Prod_Only" + ' </h4></th></tr><tr><td>'
                    # + res.to_html(index=False, border=2, justify="center") + '</td></tr>'
                    + '<tbody>'
                    + '</table>'
                    + '</center>')